import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, colors
from datetime import datetime
import os

# Define the data structure/column headers
COLUMN_HEADERS = [
    "Asset", 
    "System", 
    "Description", 
    "Quantity", 
    "Brand/Model", 
    "Comments", 
    "Photos Attached"
]

def create_report_workbook(output_dir, visit_info, processed_items):
    """
    Creates a new Excel workbook, writes header info, and populates the items list.
    
    Args:
        output_dir (str): The directory where the file will be saved.
        visit_info (dict): Data from the 'Visit Info' tab of the form.
        processed_items (list): List of dictionaries for the report items.

    Returns:
        tuple: (excel_path, excel_filename)
    """
    # 1. Setup
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Site Visit Report"
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    building_name = visit_info.get('building_name', 'Report').replace(' ', '_')
    excel_filename = f"{building_name}_Report_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)
    
    # Define styles
    bold_font = Font(bold=True)
    # Using hex color for white text on dark background
    header_font = Font(bold=True, size=10, color=colors.WHITE) 

    # 2. Write Visit Info (Header Section)
    sheet['A1'] = "SITE VISIT REPORT"
    sheet['A1'].font = Font(bold=True, size=14)
    
    # Write key-value pairs for visit info
    header_data = {
        'A3': 'Building Name:', 'B3': visit_info.get('building_name', 'N/A'),
        'A4': 'Address:', 'B4': visit_info.get('site_address', 'N/A'), # Corrected key: site_address vs building_address
        'A5': 'Technician:', 'B5': visit_info.get('technician_name', 'N/A'),
        'A6': 'Date Generated:', 'B6': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    for cell, value in header_data.items():
        sheet[cell] = value
        if cell.startswith('A'):
            sheet[cell].font = bold_font

    # Add a spacer row
    REPORT_START_ROW = 8
    sheet[f'A{REPORT_START_ROW-1}'] = "REPORTED ITEMS LIST"
    sheet[f'A{REPORT_START_ROW-1}'].font = bold_font
    
    # 3. Write Column Headers
    header_row = REPORT_START_ROW
    for col_index, header_title in enumerate(COLUMN_HEADERS):
        col_letter = openpyxl.utils.get_column_letter(col_index + 1)
        cell = sheet[f'{col_letter}{header_row}']
        cell.value = header_title
        
        # Apply style to header row
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        # Dark gray background color
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    # 4. Write Report Items List
    current_row = header_row + 1
    
    for item in processed_items:
        data_to_write = [
            item.get('asset', ''),
            item.get('system', ''),
            item.get('description', ''),
            item.get('quantity', 1),
            item.get('brand', 'N/A'),
            item.get('comments', 'N/A'),
            # CRITICAL FIX: Use 'photos' key from the client payload instead of 'image_paths'
            len(item.get('photos', [])) 
        ]
        
        for col_index, value in enumerate(data_to_write):
            col_letter = openpyxl.utils.get_column_letter(col_index + 1)
            sheet[f'{col_letter}{current_row}'] = value
            
        current_row += 1
            
    # 5. Auto-fit columns for readability (optional)
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                # Calculate max length based on content
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        # Set max width to 50 for large comments/descriptions
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50 

    # 6. Save the new workbook
    workbook.save(excel_path)
    
    return excel_path, excel_filename