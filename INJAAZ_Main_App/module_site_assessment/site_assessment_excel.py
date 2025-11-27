import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# --- Configuration for Excel Styles ---
HEADER_FILL = PatternFill(start_color='125435', end_color='125435', fill_type='solid') # INJAAZ Green
SECTION_FILL = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid') # Light Grey
TEXT_COLOR_WHITE = Font(color='FFFFFF', bold=True)
TEXT_COLOR_BLACK = Font(color='000000', bold=False)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Define the structure and display names based on your HTML form
DATA_MAPPING = {
    "Project & Client Details": {
        "client_name": "Client Name",
        "project_name": "Project Name",
        "site_address": "Site Address",
        "date_of_visit": "Date of Visit",
        "key_person_name": "Key Person/Contact Name",
        "contact_number": "Contact Number",
    },
    "Site Count & Operations": {
        "room_count": "Room Count (Units)",
        "current_team_size": "Current Team Size (Count)",
        "lift_count_total": "Total Lift Count",
        "current_team_desc": "Current Team Description/Notes",
    },
    "Facility Checklist": {
        "facility_floor": "Floors",
        "facility_ground_parking": "Ground/Parking Area",
        "facility_basement": "Basement",
        "facility_podium": "Podium",
        "facility_gym_room": "Gym/Fitness Room",
        "facility_washroom_male": "Male Washroom",
        "facility_washroom_female": "Female Washroom",
        "facility_changing_room": "Changing/Locker Room",
        "facility_play_kids_place": "Kids Play Area",
        "facility_garbage_room": "Garbage/Waste Room",
    },
    "Maintenance & Equipment Assessment": {
        "facility_equipment_condition": "Equipment Condition (General)",
        "facility_maintenance_notes": "General Maintenance Notes (Water leaks, damage, etc.)",
        "facility_equipment_notes": "Equipment Specific Notes",
    }
}

def format_value(key, value):
    """Formats the value for display in the Excel sheet."""
    # Convert 'true'/'false' strings from form to 'Yes'/'No'
    if isinstance(value, str):
        if value.lower() == 'true':
            return 'Yes'
        if value.lower() == 'false':
            return 'No'
    # Format date
    if key == 'date_of_visit' and value:
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%d %b %Y')
        except ValueError:
            return value # Return raw if date format is unexpected
    # For numeric fields, ensure they are numbers or default to N/A
    if key in ['room_count', 'current_team_size', 'lift_count_total']:
        try:
            return int(value) if value else 'N/A'
        except ValueError:
            return value
            
    # Default for all other text/notes fields
    return value if value else 'N/A'

def generate_assessment_excel(assessment_info: dict):
    """
    Converts the site assessment dictionary into a structured Excel file (XLSX).
    
    Returns a BytesIO stream and the suggested filename.
    """
    
    # 1. Prepare data for DataFrame
    excel_data = []
    
    for section_name, fields in DATA_MAPPING.items():
        # Add a section header row
        excel_data.append({'Section': section_name, 'Value': 'SECTION_HEADER'})
        
        for key, display_name in fields.items():
            value = assessment_info.get(key, '')
            formatted_value = format_value(key, value)
            
            # Use 'Section' column for the display name, and 'Value' for the data
            excel_data.append({'Section': display_name, 'Value': formatted_value})

    df = pd.DataFrame(excel_data)

    # 2. Setup Excel Writer and Styles
    excel_stream = BytesIO()
    writer = pd.ExcelWriter(excel_stream, engine='openpyxl')
    
    # Write the DataFrame to a sheet
    sheet_name = 'Site Assessment Data'
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1) 
    
    worksheet = writer.sheets[sheet_name]

    # Set column widths
    worksheet.column_dimensions['A'].width = 35  # Key/Section Name Column
    worksheet.column_dimensions['B'].width = 65  # Value Column

    # 3. Apply Professional Formatting
    
    # Header Row (A1 and B1)
    worksheet['A1'].value = 'Assessment Area'
    worksheet['B1'].value = 'Details/Findings'
    for col in ['A', 'B']:
        header_cell = worksheet[f'{col}1']
        header_cell.font = TEXT_COLOR_WHITE
        header_cell.fill = HEADER_FILL
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = thin_border
        worksheet.row_dimensions[1].height = 20 # Standard height

    row_num = 2 # Start from the first data row
    grey_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')

    for index, row_data in df.iterrows():
        is_header = row_data['Value'] == 'SECTION_HEADER'
        
        if is_header:
            # Section Header Row styling (Merge and Center)
            worksheet.merge_cells(f'A{row_num}:B{row_num}')
            header_cell = worksheet[f'A{row_num}']
            header_cell.value = row_data['Section'].upper()
            header_cell.font = TEXT_COLOR_WHITE
            header_cell.fill = HEADER_FILL
            worksheet.row_dimensions[row_num].height = 20
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = thin_border
            # Ensure the merged range also gets borders
            worksheet[f'B{row_num}'].border = thin_border 
            
        else:
            # Regular Data Row styling (Banding and Content)
            current_fill = grey_fill if (row_num - 1) % 2 == 0 else PatternFill(fill_type=None) # White fill

            # Apply styling to Column A (Key/Display Name)
            cell_a = worksheet[f'A{row_num}']
            cell_a.fill = current_fill
            cell_a.border = thin_border
            cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell_a.font = Font(bold=True) # Make the label bold
            
            # Apply styling to Column B (Value/Data)
            cell_b = worksheet[f'B{row_num}']
            cell_b.fill = current_fill
            cell_b.border = thin_border
            cell_b.alignment = wrap_alignment
            
            # Set row height dynamically for wrapped text 
            # Check length of the text in column B to determine row height
            value_str = str(row_data['Value']) if row_data['Value'] is not None else ''
            if len(value_str) > 60 and '\n' in value_str:
                 worksheet.row_dimensions[row_num].height = 75 # For multi-line text areas
            elif len(value_str) > 50:
                 worksheet.row_dimensions[row_num].height = 40 
            else:
                 worksheet.row_dimensions[row_num].height = 20 # Standard height

        row_num += 1

    # 4. Save and return
    writer.close() # CRITICAL: This finalizes the Excel file
    excel_stream.seek(0) # CRITICAL: This allows Flask's send_file to read the content
    
    client_name = assessment_info.get('client_name', 'Client').replace(' ', '_')
    date_of_visit = assessment_info.get('date_of_visit', datetime.now().strftime('%Y%m%d'))
    excel_filename = f"Site_Assessment_{client_name}_{date_of_visit}.xlsx"
    
    return excel_stream, excel_filename

# Note: The code assumes all necessary libraries (pandas, openpyxl) are installed in your environment.